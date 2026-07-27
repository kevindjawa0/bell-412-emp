from __future__ import annotations

import argparse
import json
import math
import re
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parent
DATA_DIR = PROJECT_DIR / "data"
WORKBOOK_PATH = DATA_DIR / "MAINTENANCE PROGRAM BELL 412 R01.xlsx"
OUTPUT_JSON = PROJECT_DIR / "basic_inspection_data.json"
OUTPUT_JS = PROJECT_DIR / "basic_inspection_data.js"

AIRCRAFT_PROGRAMS = {
    "OCA": {"registration": "PK-OCA", "model": "BELL 412 SP"},
    "OCD": {"registration": "PK-OCD", "model": "BELL 412 EP"},
}

PARENT_ORDER = [
    "B412-MPDI-05-25B",
    "B412-MPDI-05-100B",
    "B412-MPDI-05-300B",
    "B412-MPDI-05-600B",
    "B412-MPDI-05-5000B",
    "B412-MPDI-05-CSSD-25",
    "B412-MPDI-05-CSSD-100",
]

DISPLAY_NAMES = {
    "B412-MPDI-05-25B": "25-hour",
    "B412-MPDI-05-100B": "100-hour",
    "B412-MPDI-05-300B": "300-hour",
    "B412-MPDI-05-600B": "600-hour",
    "B412-MPDI-05-5000B": "5000-hour",
    "B412-MPDI-05-CSSD-25": "CSSD 25-hour",
    "B412-MPDI-05-CSSD-100": "CSSD 100-hour",
}

AVERAGE_MAN_HOURS_PER_TASK = {
    # Fallback assumptions for packages without confirmed Man Hours.
    "B412-MPDI-05-25B": 0.16,
    "B412-MPDI-05-100B": 0.52,
    "B412-MPDI-05-300B": 0.38,
    "B412-MPDI-05-600B": 1.08,
    "B412-MPDI-05-5000B": 2.35,
    "B412-MPDI-05-CSSD-25": 0.22,
    "B412-MPDI-05-CSSD-100": 0.34,
}

CONFIRMED_PACKAGE_MAN_HOURS = {
    "B412-MPDI-05-25B": 6.0,
    "B412-MPDI-05-100B": 24.0,
    "B412-MPDI-05-300B": 50.0,
    "B412-MPDI-05-600B": 110.0,
    "B412-MPDI-05-5000B": 700.0,
}


def round_to_half(value: float) -> float:
    return math.floor(value * 2 + 0.5) / 2


def clean_note(value: object) -> str:
    note = str(value or "").strip()
    normalized = {
        "PACAKGE": "Package",
        "PACKAGE": "Package",
        "PROGRESIVE LM/HANGAR": "Progressive LM/Hangar",
        "PROGRESSIVE LM/HANGAR": "Progressive LM/Hangar",
    }
    return normalized.get(note.upper(), note.title() if note else "Package")


def short_parent(task_card: str) -> str:
    return task_card.replace("B412-", "", 1)


def normalize_header(value: object) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def find_column_index(worksheet, labels: set[str], default_index: int, max_header_row: int = 3) -> int:
    normalized_labels = {normalize_header(label) for label in labels}

    for row in worksheet.iter_rows(min_row=1, max_row=max_header_row, values_only=True):
        for index, value in enumerate(row):
            if normalize_header(value) in normalized_labels:
                return index

    return default_index


def parse_applicability(value: object) -> set[str]:
    text = str(value or "").upper()
    return {match.group(0) for match in re.finditer(r"\b(?:OCA|OCD)\b", text)}


def format_applicability(keys: set[str]) -> str:
    ordered_keys = [key for key in AIRCRAFT_PROGRAMS if key in keys]
    return " | ".join(ordered_keys) if ordered_keys else "OCA | OCD"


def workbook_name(workbook_path: Path | str | BinaryIO) -> str:
    if isinstance(workbook_path, (str, Path)):
        return Path(workbook_path).name

    return str(getattr(workbook_path, "name", "uploaded workbook"))


def read_mp_metadata(workbook) -> dict[str, dict[str, object]]:
    worksheet = workbook["MP"]
    applicability_index = find_column_index(worksheet, {"APPLICABILITY", "APL"}, 12)
    metadata = {}

    for row in worksheet.iter_rows(min_row=4, values_only=True):
        task_card = row[2]
        if not task_card:
            continue
        applicability_keys = parse_applicability(row[applicability_index])

        metadata[str(task_card)] = {
            "intervalText": str(row[5] or "").strip(),
            "intervalFlightHours": row[7],
            "calendarLimitDays": row[9],
            "applicabilityText": str(row[applicability_index] or "").strip(),
            "applicabilityKeys": applicability_keys,
            "taskTitle": str(row[15] or "").strip(),
            "taskDescription": str(row[16] or "").strip(),
        }

    return metadata


def read_summary_notes(workbook) -> dict[str, str]:
    worksheet = workbook["Parent Tasks"]
    notes = {}

    for row in worksheet.iter_rows(min_row=3, max_row=9, values_only=True):
        task_card = row[5]
        if not task_card:
            continue
        notes[str(task_card)] = clean_note(row[8])

    return notes


def count_child_tasks(workbook, metadata: dict[str, dict[str, object]]) -> dict[str, dict[str, int]]:
    worksheet = workbook["Parent Tasks"]
    applicability_index = find_column_index(worksheet, {"APPLICABILITY", "APL"}, 15, max_header_row=13)
    counts = {
        short_parent(task_card): {"all": 0, **{key: 0 for key in AIRCRAFT_PROGRAMS}}
        for task_card in PARENT_ORDER
    }

    for row in worksheet.iter_rows(min_row=14, values_only=True):
        row_type = row[1]
        parent = row[3]
        if row_type != "CHILD TASK" or parent not in counts:
            continue

        task_card = str(row[5] or "")
        applicability_keys = parse_applicability(row[applicability_index])
        if not applicability_keys:
            applicability_keys = set(metadata.get(task_card, {}).get("applicabilityKeys", set()))
        if not applicability_keys:
            applicability_keys = set(AIRCRAFT_PROGRAMS)

        counts[str(parent)]["all"] += 1
        for aircraft_key in applicability_keys:
            if aircraft_key in AIRCRAFT_PROGRAMS:
                counts[str(parent)][aircraft_key] += 1

    return counts


def calculate_package_man_hours(task_card: str, children: int, total_children: int) -> float:
    if task_card in CONFIRMED_PACKAGE_MAN_HOURS:
        if not total_children:
            return 0
        ratio = children / total_children
        return round(CONFIRMED_PACKAGE_MAN_HOURS[task_card] * ratio, 1)

    return round_to_half(children * AVERAGE_MAN_HOURS_PER_TASK[task_card])


def build_aircraft_parent_tasks(
    aircraft_key: str,
    metadata: dict[str, dict[str, object]],
    notes: dict[str, str],
    child_counts: dict[str, dict[str, int]],
) -> list[dict[str, object]]:
    parent_tasks = []

    for task_card in PARENT_ORDER:
        meta = metadata.get(task_card, {})
        parent_key = short_parent(task_card)
        counts = child_counts.get(parent_key, {})
        children = counts.get(aircraft_key, 0)
        total_children = counts.get("all", children)

        if children == 0:
            continue

        man_hours = calculate_package_man_hours(task_card, children, total_children)
        avg_hours = man_hours / children if children else 0
        applicability_keys = set(meta.get("applicabilityKeys") or AIRCRAFT_PROGRAMS)

        parent_tasks.append(
            {
                "parentPackage": DISPLAY_NAMES[task_card],
                "taskCard": task_card,
                "intervalText": meta.get("intervalText", ""),
                "intervalFlightHours": meta.get("intervalFlightHours"),
                "calendarLimitDays": meta.get("calendarLimitDays"),
                "childTasks": children,
                "totalChildTasks": total_children,
                "averageManHoursPerTask": round(avg_hours, 4),
                "manHours": man_hours,
                "currentNote": notes.get(task_card, "Package"),
                "taskTitle": meta.get("taskTitle", ""),
                "applicability": sorted(applicability_keys),
                "applicabilityLabel": format_applicability(applicability_keys),
            }
        )

    return parent_tasks


def summarize_parent_tasks(parent_tasks: list[dict[str, object]]) -> dict[str, object]:
    return {
        "childTasks": sum(int(item["childTasks"]) for item in parent_tasks),
        "manHours": round_to_half(sum(float(item["manHours"]) for item in parent_tasks)),
    }

def build_basic_inspection_data(workbook_path: Path | str | BinaryIO = WORKBOOK_PATH) -> dict[str, object]:
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    metadata = read_mp_metadata(workbook)
    notes = read_summary_notes(workbook)
    child_counts = count_child_tasks(workbook, metadata)
    aircraft_programs = []

    for aircraft_key, aircraft in AIRCRAFT_PROGRAMS.items():
        parent_tasks = build_aircraft_parent_tasks(aircraft_key, metadata, notes, child_counts)
        aircraft_programs.append(
            {
                "key": aircraft_key,
                "registration": aircraft["registration"],
                "model": aircraft["model"],
                "parentTasks": parent_tasks,
                "totals": summarize_parent_tasks(parent_tasks),
            }
        )

    default_parent_tasks = aircraft_programs[0]["parentTasks"] if aircraft_programs else []

    return {
        "model": "BELL 412",
        "sourceWorkbook": workbook_name(workbook_path),
        "methodology": (
            "Primary inspection Man Hours use the confirmed package totals supplied for the 25-hour, "
            "100-hour, 300-hour, 600-hour, and 5000-hour blocks. Corrosion-control CSSD rows retain "
            "the previous estimating assumption until confirmed Man Hours are supplied. Applicability "
            "is read from the workbook so OCA and OCD maintenance programs are modeled separately."
        ),
        "aircraftPrograms": aircraft_programs,
        "parentTasks": default_parent_tasks,
        "totals": summarize_parent_tasks(default_parent_tasks),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build basic inspection dashboard data.")
    parser.add_argument(
        "--source",
        type=Path,
        default=WORKBOOK_PATH,
        help="Path to the maintenance program Excel workbook. Defaults to data/MAINTENANCE PROGRAM BELL 412 R01.xlsx.",
    )
    args = parser.parse_args()

    if not args.source.exists():
        raise FileNotFoundError(
            f"Workbook not found: {args.source}. Upload it in Streamlit, or place it under the local data/ "
            "folder for offline regeneration. Do not commit confidential workbooks."
        )

    data = build_basic_inspection_data(args.source)
    OUTPUT_JSON.write_text(json.dumps(data, indent=2), encoding="utf-8")
    OUTPUT_JS.write_text(
        "window.BASIC_INSPECTION_DATA = "
        + json.dumps(data, indent=2)
        + ";\n",
        encoding="utf-8",
    )

    print(f"Wrote {OUTPUT_JSON.name} and {OUTPUT_JS.name}")
    for aircraft in data["aircraftPrograms"]:
        print(f"{aircraft['model']} ({aircraft['registration']})")
        for item in aircraft["parentTasks"]:
            print(
                f"  {item['parentPackage']}: {item['childTasks']} tasks, "
                f"{item['manHours']:.1f} Man Hours"
            )


if __name__ == "__main__":
    main()
